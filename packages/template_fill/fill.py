"""テンプレート Excel への値の流し込み。"""

from collections.abc import Mapping
from pathlib import Path

from openpyxl import load_workbook

from .mapping import TemplateMapping


class FillError(ValueError):
    """入力データとマッピングの不整合、テンプレート不備の場合に送出する。"""


def fill_template(
    mapping: TemplateMapping,
    data: Mapping[str, object],
    output_path: Path,
) -> Path:
    """マッピングに従いテンプレートへ data を流し込み、output_path に xlsx を保存する。

    マッピング側の全フィールドが必須。data 側の未知キーも typo 検出のためエラーにする。
    """
    unknown = sorted(set(data) - set(mapping.fields))
    if unknown:
        raise FillError(f"マッピングに存在しないフィールドです: {', '.join(unknown)}")
    missing = sorted(set(mapping.fields) - set(data))
    if missing:
        raise FillError(f"必須フィールドが入力にありません: {', '.join(missing)}")

    if not mapping.template.is_file():
        raise FillError(f"テンプレートが見つかりません: {mapping.template}")

    workbook = load_workbook(mapping.template)
    for name, ref in mapping.fields.items():
        if ref.sheet is None:
            sheet = workbook.active
            if sheet is None:
                raise FillError(f"アクティブシートがありません: {mapping.template}")
        elif ref.sheet in workbook.sheetnames:
            sheet = workbook[ref.sheet]
        else:
            raise FillError(
                f"フィールド '{name}': シート '{ref.sheet}' がテンプレートにありません"
            )
        sheet[ref.cell] = data[name]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
