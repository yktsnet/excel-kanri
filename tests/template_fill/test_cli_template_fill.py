import io
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from packages.template_fill import cli as cli_module


@pytest.fixture
def mapping_file(tmp_path: Path) -> Path:
    workbook = Workbook()
    workbook.save(tmp_path / "template.xlsx")
    path = tmp_path / "mapping.yaml"
    path.write_text("template: template.xlsx\nfields:\n  name: B3\n", encoding="utf-8")
    return path


def test_main_writes_output_and_prints_path(
    tmp_path: Path, mapping_file: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    data = tmp_path / "data.json"
    data.write_text(json.dumps({"name": "山田 太郎"}), encoding="utf-8")
    output = tmp_path / "out.xlsx"

    exit_code = cli_module.main([str(mapping_file), str(data), "-o", str(output)])

    assert exit_code == 0
    assert capsys.readouterr().out.strip() == str(output)
    assert load_workbook(output).active["B3"].value == "山田 太郎"


def test_main_reads_data_from_stdin(
    tmp_path: Path,
    mapping_file: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("sys.stdin", io.StringIO(json.dumps({"name": "x"})))
    output = tmp_path / "out.xlsx"

    exit_code = cli_module.main([str(mapping_file), "-", "-o", str(output)])

    assert exit_code == 0
    assert output.exists()


def test_main_rejects_non_object_data(
    tmp_path: Path, mapping_file: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    data = tmp_path / "data.json"
    data.write_text("[1, 2, 3]", encoding="utf-8")

    exit_code = cli_module.main([str(mapping_file), str(data), "-o", str(tmp_path / "o.xlsx")])

    assert exit_code == 2
    assert "ルートはオブジェクト" in capsys.readouterr().err


def test_main_reports_fill_error_and_returns_2(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    mapping = tmp_path / "mapping.yaml"
    mapping.write_text("template: missing.xlsx\nfields:\n  name: B3\n", encoding="utf-8")
    data = tmp_path / "data.json"
    data.write_text("{}", encoding="utf-8")

    exit_code = cli_module.main([str(mapping), str(data), "-o", str(tmp_path / "o.xlsx")])

    assert exit_code == 2
    assert capsys.readouterr().err.startswith("error:")
