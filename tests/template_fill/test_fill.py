from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from packages.template_fill import (
    CellRef,
    FillError,
    TemplateMapping,
    fill_template,
)


@pytest.fixture
def template(tmp_path: Path) -> Path:
    workbook = Workbook()
    workbook.create_sheet("Detail")
    path = tmp_path / "template.xlsx"
    workbook.save(path)
    return path


def make_mapping(template: Path, fields: dict[str, CellRef]) -> TemplateMapping:
    return TemplateMapping(template=template, fields=fields)


def test_fill_template_writes_values(template: Path, tmp_path: Path) -> None:
    mapping = make_mapping(
        template,
        {
            "name": CellRef(sheet=None, cell="B3"),
            "room": CellRef(sheet="Detail", cell="D5"),
        },
    )
    output = tmp_path / "out" / "filled.xlsx"
    result = fill_template(mapping, {"name": "山田 太郎", "room": 101}, output)

    assert result == output
    workbook = load_workbook(output)
    active = workbook.active
    assert active is not None
    assert active["B3"].value == "山田 太郎"
    assert workbook["Detail"]["D5"].value == 101


def test_fill_template_rejects_unknown_field(template: Path, tmp_path: Path) -> None:
    mapping = make_mapping(template, {"name": CellRef(sheet=None, cell="B3")})
    with pytest.raises(FillError, match="マッピングに存在しない"):
        fill_template(mapping, {"name": "x", "nmae": "y"}, tmp_path / "o.xlsx")


def test_fill_template_rejects_missing_field(template: Path, tmp_path: Path) -> None:
    mapping = make_mapping(
        template,
        {"name": CellRef(sheet=None, cell="B3"), "room": CellRef(sheet=None, cell="C1")},
    )
    with pytest.raises(FillError, match="必須フィールド"):
        fill_template(mapping, {"name": "x"}, tmp_path / "o.xlsx")


def test_fill_template_rejects_unknown_sheet(template: Path, tmp_path: Path) -> None:
    mapping = make_mapping(template, {"name": CellRef(sheet="Nope", cell="B3")})
    with pytest.raises(FillError, match="シート 'Nope'"):
        fill_template(mapping, {"name": "x"}, tmp_path / "o.xlsx")


def test_fill_template_rejects_missing_template(tmp_path: Path) -> None:
    mapping = make_mapping(tmp_path / "none.xlsx", {"name": CellRef(sheet=None, cell="B3")})
    with pytest.raises(FillError, match="テンプレートが見つかりません"):
        fill_template(mapping, {"name": "x"}, tmp_path / "o.xlsx")
